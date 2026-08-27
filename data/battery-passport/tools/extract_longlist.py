# -*- coding: utf-8 -*-
"""Derive a requirement index from the Battery Pass data attribute longlist (XLSX).

The longlist is a spreadsheet: one row per data attribute, with a column per
battery category saying whether that attribute is required for it, a column
naming the legal provision behind it, and columns for access rights, unit, data
format and granularity. This reads those columns and writes them out one record
per row.

Two things it deliberately does not do:

  * It does not decide. The sheet's own legend distinguishes "x" (required by the
    Batteries Regulation) from "(x)" (required by other instruments) from "o"
    (voluntary); a blank is left as unclear rather than read as optional.
  * It does not copy the column that restates a separate specification's
    requirements. That column belongs to a document with different reuse terms,
    so only its presence is recorded, along with the chapter pointer the sheet
    gives. The pointer is a citation; the wording stays where it was published.

`--public` selects the redistribution profile: the set of fields left out of an
index meant to be published. **That set is currently empty**, and the rule that
keeps it empty is worth stating, because it was got wrong once in the other
direction: *what a document publishes in its own column is that document's*. The
chapter number is a value this spreadsheet prints itself, under a licence that
permits reuse; it points at another document without quoting a word of it. Only
wording lifted out of that other document would belong in the profile.

The flag stays because the distinction is real and may bite again, and because a
published index must be something this tool produces rather than something
someone edited afterwards. Both variants are reproducible from the same source
bytes; neither is a hand-trimmed copy of the other.

The workbook is opened read-only with links disabled: a spreadsheet is data
here, never something to execute.

Usage:
    python3 tools/extract_longlist.py \
        --xlsx "sources/batterypass/2026_BatteryPass-Ready_DataAttributeLongList_v1.3.xlsx" \
        --out requirements-longlist.json
"""

import argparse
import os
import re
import sys
import warnings

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import _common  # noqa: E402

SOURCE = "longlist"
SHEET = "Data attribute longlist_DR_v1.3"
VERSION = "v1.3"
FIRST_DATA_ROW = 8

# 1-based column numbers, as the sheet lays them out.
COL = {
    "number": 2,
    "spec_chapter": 3,
    "ev": 4,
    "lmt": 5,
    "industrial_other": 6,
    "industrial_stationary": 7,
    "category": 8,
    "sub_category": 9,
    "attribute": 10,
    "definition": 11,
    "requirement_regulation": 12,
    "requirement_spec": 13,  # separate specification -- presence only, see above
    "legal_reference": 14,
    "unit": 15,
    "data_format": 16,
    "access": 17,
    "static_or_dynamic": 18,
    "update_requirement": 19,
    "granularity": 20,
}
CATEGORY_COLUMNS = (
    ("EV", "ev"),
    ("LMT", "lmt"),
    ("industrial-other-above-2kWh", "industrial_other"),
    ("industrial-stationary-above-2kWh", "industrial_stationary"),
)

# The sheet's own legend, row 6: "x = mandatory per BattReg; (x) = mandatory per
# ESPR / JTC-24; o = voluntary". A blank cell carries no statement at all.
MARK_MEANING = {
    "x": "required-by-batteries-regulation",
    "(x)": "required-by-other-instrument",
    "o": "voluntary",
    "-": "not-applicable",
    "": "not-stated",
}

# Fields left out of the redistribution profile (--public). Empty, and the
# reason is the rule: a value the sheet prints in its own column is the sheet's
# to publish. Only wording lifted from a document published under other terms
# would be listed here.
PUBLIC_PROFILE_OMITS = ()

ACCESS_MAP = {
    "public": "public",
    "persons with a legitimate interest": "legitimate-interest",
    "persons with a legitimate interest and the commission": "legitimate-interest",
    "notified bodies, market surveillance authorities and the commission": "authorities",
}


def cell(row, name):
    value = row[COL[name] - 1]
    if value is None:
        return ""
    return _common.squeeze(str(value))


def split_references(text):
    """The legal reference cell holds one or more provisions, separated by
    line breaks, semicolons or commas between citations."""
    parts = re.split(r"[\n;]+", text or "")
    return [_common.squeeze(p).rstrip(",") for p in parts if _common.squeeze(p)]


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out", default="requirements-longlist.json")
    ap.add_argument(
        "--public",
        action="store_true",
        help=(
            "produce the redistribution profile of this index. It currently "
            "omits nothing: a value the sheet prints in its own column is the "
            "sheet's to publish"
        ),
    )
    args = ap.parse_args(argv)

    with warnings.catch_warnings():
        # openpyxl warns about spreadsheet extensions it drops; dropping them is
        # what read-only parsing is for.
        warnings.simplefilter("ignore")
        book = openpyxl.load_workbook(
            args.xlsx, read_only=True, data_only=True, keep_links=False
        )
    if SHEET not in book.sheetnames:
        raise SystemExit(
            "sheet %r not in %r -- the workbook edition changed" % (SHEET, book.sheetnames)
        )
    sheet = book[SHEET]
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        rows = [r for r in sheet.iter_rows(min_row=FIRST_DATA_ROW, values_only=True)]

    records, notes = [], []
    marks_seen = {}
    variant = "public" if args.public else "full"
    for row in rows:
        number = cell(row, "number")
        attribute = cell(row, "attribute")
        if not number and not attribute:
            continue

        marks, applies_to = {}, []
        for label, key in CATEGORY_COLUMNS:
            raw = cell(row, key).replace(" ", "")
            meaning = MARK_MEANING.get(raw)
            if meaning is None:
                meaning = "unrecognised:%s" % raw
                notes.append("row %s: unrecognised applicability mark %r" % (number, raw))
            marks[label] = meaning
            marks_seen[meaning] = marks_seen.get(meaning, 0) + 1
            if meaning in ("required-by-batteries-regulation", "required-by-other-instrument"):
                applies_to.append(label)

        values = set(marks.values())
        if "required-by-batteries-regulation" in values:
            mandatory, condition = "yes", ""
        elif "required-by-other-instrument" in values:
            mandatory = "yes"
            condition = (
                "the sheet marks this as required by another instrument "
                "(its legend: ESPR / JTC-24), not by the Batteries Regulation"
            )
        elif values == {"voluntary"} or values <= {"voluntary", "not-applicable", "not-stated"}:
            mandatory = "no" if "voluntary" in values else "unclear"
            condition = ""
        else:
            mandatory, condition = "unclear", ""

        access_raw = cell(row, "access")
        access = ACCESS_MAP.get(access_raw.lower(), "n/a")
        if access == "n/a" and access_raw:
            notes.append("row %s: unrecognised access wording %r" % (number, access_raw))

        spec_chapter = cell(row, "spec_chapter").strip(" -") or ""
        pointers = {
            "separate_spec_chapter": spec_chapter,
            "separate_spec_has_requirement_text": bool(cell(row, "requirement_spec")),
        }
        for field in PUBLIC_PROFILE_OMITS if args.public else ():
            pointers.pop(field, None)
        records.append(
            _common.record(
                id="%s:%s" % (SOURCE, number),
                kind="attribute",
                section="longlist %s row %s" % (VERSION, number),
                text=attribute,
                mandatory=mandatory,
                condition=condition,
                access=access,
                applies_to=applies_to,
                joins=[],
                number=number,
                category=cell(row, "category"),
                sub_category=cell(row, "sub_category"),
                definition=cell(row, "definition"),
                requirement_regulation=cell(row, "requirement_regulation"),
                legal_reference=cell(row, "legal_reference"),
                legal_references=split_references(cell(row, "legal_reference")),
                applicability=marks,
                access_as_written=access_raw,
                unit=cell(row, "unit"),
                data_format=cell(row, "data_format"),
                static_or_dynamic=cell(row, "static_or_dynamic"),
                update_requirement=cell(row, "update_requirement"),
                granularity=cell(row, "granularity"),
                **pointers
            )
        )
    book.close()

    without_legal_reference = [r["id"] for r in records if not r["legal_references"]]
    if without_legal_reference:
        notes.append(
            "%d rows cite no legal provision, so they cannot be joined to the "
            "regulation by citation alone: %s"
            % (len(without_legal_reference), ", ".join(without_legal_reference))
        )

    doc = _common.write_index(
        args.out,
        SOURCE,
        records,
        [
            {
                "file": _common.provenance_path(args.xlsx),
                "sha256": _common.sha256_file(args.xlsx),
                "version": VERSION,
                "sheet": SHEET,
            }
        ],
        counts=dict(
            {
                "by_applicability_mark": dict(sorted(marks_seen.items())),
                "by_granularity": _common.tally(records, "granularity"),
                "by_static_or_dynamic": _common.tally(records, "static_or_dynamic"),
                "variant": variant,
            },
            **(
                {}
                if "separate_spec_chapter" in PUBLIC_PROFILE_OMITS and args.public
                else {
                    "rows_with_separate_spec_chapter": sum(
                        1 for r in records if r["separate_spec_chapter"]
                    )
                }
            )
        ),
        notes=notes,
    )
    _common.report(doc, args.out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
